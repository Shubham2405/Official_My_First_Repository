
def correct_worksheet(name):
    import openpyxl as xl
    wb=xl.load_workbook(name)
    sht=wb['Sheet1']
    for x in range(2,5):
        cll=sht.cell(x,3)
        new_value=(cll.value)*(0.9)
        crtd_cell=sht.cell(x,4)
        crtd_cell.value=new_value
    return(wb)

(correct_worksheet('transactions.xlsx')).save('ttt2.xlsx')
